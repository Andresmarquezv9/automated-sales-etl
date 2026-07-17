# ==================================================================================================================================================================================================
# LIBRERIAS
# ==================================================================================================================================================================================================

import sys
import re
import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook




# ==================================================================================================================================================================================================
# VARIABLES GLOBALES
# ==================================================================================================================================================================================================

RE_DATE    = re.compile(r"^\d{2}/\d{2}/\d{2,4}$")    
RE_SECTION = re.compile(r"^[A-Z]{2}\s{2}\S+\s{2}.+") 
RE_NOISE   = re.compile(r"^[=\-]{3,}")                
SKIP_VALS  = {"date", "total", "--------", ""}



# ==================================================================================================================================================================================================
# FUNCIONES
# ==================================================================================================================================================================================================

def detect_columns(ws) -> list:
    """
    Busca la primera fila que empiece con 'Date' y 'Unit'
    y la usa como nombres de columna.
    Si hay columnas duplicadas (ej. dos 'Date') añade sufijo _2, _3, etc.
    """
    for row in ws.iter_rows(values_only=True):
        vals = [str(v or "").strip() for v in row]
        if vals[0].lower() == "date" and vals[1].lower() == "unit":
            cols, seen = [], {}
            for v in vals:
                key = v if v else "Col"
                seen[key] = seen.get(key, 0) + 1
                cols.append(f"{key}_{seen[key]}" if seen[key] > 1 else key)
            return cols
    return []

def parse_section(val: str) -> tuple:
    """
    Extrae (group, item_code, description) del encabezado de sección.
    Ejemplo: 'AA  BA001  BREAD/BAGUETTE/PLAIN |'
             → ('AA', 'BA001', 'BREAD/BAGUETTE/PLAIN')
    """
    parts = val.split()
    group = parts[0] if parts else ""
    code  = parts[1] if len(parts) > 1 else ""
    desc  = " ".join(parts[2:]).rstrip("|").strip()
    return group, code, desc

def is_noise(val: str) -> bool:
    return bool(RE_NOISE.match(val)) or val.lower() in SKIP_VALS


# ==================================================================================================================================================================================================
# RUTAS
# ==================================================================================================================================================================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

BRONZE_DIR = os.path.join(BASE_DIR, "..", "1_landing_bronze")
SILVER_DIR = os.path.join(BASE_DIR, "..", "2_silver")

INPUT_FILE = os.path.join(BRONZE_DIR, "prueba report - January 2025 to April 2026.xlsx")

excel_name = "prueba report - January 2025 to April 2026.xlsx"
name_wo_ext = excel_name.replace(".xlsx", "")
original_dates = name_wo_ext.split(" - ")[-1]  
format_dates = original_dates.replace(" to ", " - ")
parquet_name = f"Sales_table {format_dates}.parquet"


OUTPUT_FILE = os.path.join(SILVER_DIR, parquet_name)


# ==================================================================================================================================================================================================
# PROCESAMIENTO
# ==================================================================================================================================================================================================

wb = load_workbook(INPUT_FILE, read_only=True)
ws = wb.worksheets[0]
print(f"   Hoja activa : '{ws.title}'")
 
columns = detect_columns(ws)
if not columns:
    raise ValueError(
        "No se encontró la fila de encabezado con 'Date' / 'Unit'.\n"
        "Verifica que el archivo tenga el formato esperado."
    )
print(f"   Columnas    : {columns}")

records       = []
current_group = current_code = current_desc = ""
 
for row in ws.iter_rows(values_only=True):
    col0 = str(row[0] or "").strip()
 
    
    if RE_SECTION.match(col0) and row[1] is None:
        current_group, current_code, current_desc = parse_section(col0)
        continue
 
    
    if is_noise(col0):
        continue
 
   
    if RE_DATE.match(col0):
        vals = list(row)
        while len(vals) < len(columns):
            vals.append(None)
        record = {
            "Group":       current_group,
            "Item_Code":   current_code,
            "Description": current_desc,
        }
        record.update(dict(zip(columns, vals[:len(columns)])))
        records.append(record)
 
wb.close()
print(f"\n Filas de datos extraídas : {len(records):,}")

df = pd.DataFrame(records)
 

for col in [c for c in df.columns if "date" in c.lower()]:
    df[col] = pd.to_datetime(
        df[col].astype(str), dayfirst=True, errors="coerce", format="mixed"
    )
 

for col in ["Quantity", "Rate", "Value"]:
    if col in df.columns:
        df[col] = pd.to_numeric(
            df[col].astype(str)
                   .str.replace(",", "", regex=False)
                   .str.replace("$", "", regex=False)
                   .str.strip(),
            errors="coerce",
        )
 

name_col = next((c for c in df.columns if "name" in c.lower()), None)
if name_col:
    df[name_col] = (
        df.groupby(["Group", "Item_Code"])[name_col]
          .transform(lambda s: s.ffill().bfill())
    )
 

lead = ["Group", "Item_Code", "Description"]
df = (df[lead + [c for c in df.columns if c not in lead]]
        .sort_values(["Group", "Item_Code", "Date"])
        .reset_index(drop=True))


try:
    df.to_parquet(OUTPUT_FILE, index=False)
    
    with open(r"C:\Data_lake\Scripts\log.txt", "w") as f:
        f.write("¡ÉXITO! Archivo guardado correctamente en Silver.")
except Exception as e:
    
    import traceback
    with open(r"C:\Data_lake\Scripts\log.txt", "w") as f:
        f.write("ERROR AL GUARDAR:\n" + traceback.format_exc())
print(f"\n Guardado en : {OUTPUT_FILE}")
print(f"   Formato     : TXT separado por tabulaciones (\\t)")
print(f"   Filas       : {len(df):,}  |  Columnas: {len(df.columns)}")